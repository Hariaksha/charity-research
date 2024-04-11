# Program to combine spreadsheets into one.
import openpyxl

def main():

    states_list = ["AL", "AK", "AZ", "AR", "CA", "CO", "CT", "DE", "FL", "GA", "HI", "ID", "IL", "IN", "IA", "KS", "KY", "LA", "ME", "MD", "MA", "MI", "MN", "MS", "MO", "MT", "NE", "NV", "NH", "NJ", "NM", "NY", "NC", "ND", "OH", "OK", "OR", "PA", "RI", "SC", "SD", "TN", "TX", "UT", "VT", "VA", "WA", "WV", "WI", "WY"]
    new_workbook = openpyxl.load_workbook('data/april_data.xlsx')
    new_ws = new_workbook.active
    print("Opened new worbook")

    for state in states_list:
        try:
            workbook = openpyxl.load_workbook(f'data/american/states/{state}_data.xlsx')
            ws = workbook.active
            for i in range(2, ws.max_row + 1):
                if ws[f'P{i}'].value != None:
                    new_ws.append([ws[f'A{i}'].value, ws[f'B{i}'].value, ws[f'C{i}'].value, ws[f'D{i}'].value, ws[f'E{i}'].value, ws[f'F{i}'].value, ws[f'G{i}'].value, ws[f'H{i}'].value, ws[f'I{i}'].value, ws[f'J{i}'].value, ws[f'K{i}'].value, ws[f'L{i}'].value, ws[f'M{i}'].value, ws[f'N{i}'].value, ws[f'O{i}'].value, ws[f'P{i}'].value])
            print(f"Finished {state} data")
        except Exception as e:
            print(f"Encountered {e} error. Moving to next state")
            continue
    new_workbook.save("data/april_data.xlsx")
    print("saved")
    
if __name__=="__main__":
    main()