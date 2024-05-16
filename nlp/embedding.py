# https://www.geeksforgeeks.org/doc2vec-in-nlp/
from gensim.models.doc2vec import Doc2Vec, TaggedDocument
import openpyxl, datetime

# Get start time
start = datetime.datetime.now()

# Open data spreadsheet
workbook = openpyxl.load_workbook(f'data/american/data.xlsx')
ws = workbook.active
print("Opened spreadsheet")

# Populate document array with mission statements
data = []
for i in range(2, ws.max_row + 1):
    data.append(ws[f'Q{i}'].value)
print("Populated array")

# Create TaggedDocuments
tagged_data = [TaggedDocument(words=doc, tags=[str(i)]) for i, doc in enumerate(data)]
print("Create TaggedDocuments")

# Train Doc2Vec model
model = Doc2Vec(vector_size=20, min_count=2, epochs=50) # https://www.geeksforgeeks.org/epoch-in-machine-learning/
model.build_vocab(tagged_data)
model.train(tagged_data, total_examples=model.corpus_count, epochs=model.epochs)

# Get document vectors
document_vectors = [model.infer_vector(doc) for doc in data]

# Print start and end times
end = datetime.datetime.now()
print("Start:", start.strftime("%d %b, %I:%M%p"), "\nLast:", end.strftime("%d %b, %I:%M%p"))