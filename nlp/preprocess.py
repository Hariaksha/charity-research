import spacy 
from spacy.tokens import doc
import openpyxl

def main():
    nlp = spacy.load('en_core_web_lg')
    print("Loaded model")

    workbook = openpyxl.load_workbook('data/preliminary_march_data/data.xlsx')
    ws = workbook.active

    workbook2 = openpyxl.load_workbook('nlp/preprocessed-missions.xlsx')
    ws2 = workbook2.active
    print("opened workbooks")

    for i in range(2, ws.max_row + 1):
        if i == 502:
            break
        mission = str(ws[f'P{i}'].value)
        ein = ws[f'A{i}'].value
        name = ws[f'B{i}'].value
        ntee = ws[f'I{i}'].value
        asset = ws[f'K{i}'].value
        revenue = ws[f'O{i}'].value 
        doc = nlp(mission) # turn mission into a spaCy document
        lemmatized_text = " ".join([token.lemma_.lower() for token in doc if not token.is_stop and not token.is_punct])
        ws2.append([ein, name, mission, lemmatized_text, revenue, asset, ntee])

    workbook2.save('nlp/preprocessed-missions.xlsx')
    print("saved")

if __name__=="__main__":
    main()
